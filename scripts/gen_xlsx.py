#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate a structured Excel file from Huawei Cloud solution page data.

Usage:
    python gen_xlsx.py data.json output.xlsx

JSON schema matches the SKILL.md schema. Each section is one row.
Column B uses CellRichText: bold titles on their own line, normal content below.

Post-processes xlsx to add xml:space='preserve' on all <t> tags,
ensuring Feishu renders newlines in rich-text cells.
"""

import json
import sys
import zipfile
import shutil
import tempfile
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

BOLD_INLINE = InlineFont(b=True)
NORMAL_INLINE = InlineFont(b=False)
WRAP = Alignment(wrap_text=True, vertical="top")
BOLD_FONT = Font(bold=True)


def rich(parts):
    """Build CellRichText from list of (text, is_bold) tuples."""
    blocks = []
    for text, is_bold in parts:
        if text:
            blocks.append(TextBlock(BOLD_INLINE if is_bold else NORMAL_INLINE, text))
    return CellRichText(*blocks)


def build_cell(items):
    """Build rich text for a single cell from a list of items.
    Each item: {title, content?, fields?}"""
    parts = []
    for i, item in enumerate(items, 1):
        parts.append((f"{i}. {item['title']}\n", True))
        if "content" in item:
            parts.append((item["content"] + "\n", False))
        if "fields" in item:
            for key, value in item["fields"].items():
                if isinstance(value, list):
                    parts.append((f"{key}：\n", True))
                    for v in value:
                        parts.append((f"  - {v}\n", False))
                else:
                    parts.append((f"{key}：\n", True))
                    parts.append((str(value) + "\n", False))
    return rich(parts)


def fix_xml_spacepreserve(output_path):
    """Post-process xlsx to add xml:space='preserve' on all <t> tags."""
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    with zipfile.ZipFile(output_path, 'r') as zin:
        with zipfile.ZipFile(tmp.name, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == 'xl/worksheets/sheet1.xml':
                    text = data.decode('utf-8')
                    text = re.sub(
                        r'<t(?![^>]*xml:space)([^>]*)>',
                        r'<t\1 xml:space="preserve">',
                        text
                    )
                    data = text.encode('utf-8')
                zout.writestr(item, data)
    shutil.move(tmp.name, output_path)


def write_section(ws, name, data_value):
    """Write a section row. data_value is a list of items or a plain string."""
    ws.append([name, ""])
    row = ws.max_row
    cell = ws.cell(row=row, column=2)
    if isinstance(data_value, list):
        cell.value = build_cell(data_value)
    else:
        cell.value = str(data_value)
    cell.alignment = WRAP


def generate(data, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = data.get("title", "Solution")[:31]

    # Header row
    ws.append(["项目", "内容"])
    for col in [1, 2]:
        ws.cell(row=1, column=col).font = BOLD_FONT

    # --- Core fields (from detail page) ---
    write_section(ws, "标题", data.get("title", ""))
    write_section(ws, "简介", data.get("intro", ""))
    write_section(ws, "适用客户",
                  "\n".join(f"{i}. {c}" for i, c in enumerate(data.get("customers", []), 1)))
    write_section(ws, "方案优势", data.get("advantages", []))
    write_section(ws, "架构与部署", data.get("architecture", []))
    write_section(ws, "应用场景", data.get("scenarios", []))
    write_section(ws, "解决方案实践拓展", data.get("extensions", []))

    # --- Support doc fields ---
    if data.get("preparation"):
        write_section(ws, "准备工作", data["preparation"])
    if data.get("quick_deploy"):
        write_section(ws, "快速部署", data["quick_deploy"])
    if data.get("getting_started"):
        write_section(ws, "开始使用", data["getting_started"])
    if data.get("quick_uninstall"):
        write_section(ws, "快速卸载", data["quick_uninstall"])
    if data.get("cost_planning"):
        write_section(ws, "资源和成本规划", data["cost_planning"])
    if data.get("optimization"):
        write_section(ws, "优化建议", data["optimization"])

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 100

    wb.save(output_path)
    fix_xml_spacepreserve(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    data_path = sys.argv[1] if len(sys.argv) > 1 else "data.json"
    output_path = sys.argv[2] if len(sys.argv) > 2 else "output.xlsx"
    with open(data_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    generate(data, output_path)
